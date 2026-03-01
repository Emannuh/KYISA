# admin_dashboard/export_views.py — PDF & Excel export for system activity logs
import io
from datetime import datetime, timedelta

from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Q
from django.contrib.auth import get_user_model

from .models import ActivityLog

User = get_user_model()


def superadmin_required(user):
    return user.is_superuser or user.is_staff


def _filtered_logs(request):
    """Apply the same filters used by the activity_logs list view."""
    action_filter = request.GET.get("action", "")
    user_filter = request.GET.get("user", "")
    date_from = request.GET.get("date_from", "")
    date_to = request.GET.get("date_to", "")
    search_query = request.GET.get("search", "")

    logs = ActivityLog.objects.select_related("user", "content_type").all()

    if action_filter:
        logs = logs.filter(action=action_filter)
    if user_filter:
        logs = logs.filter(user_id=user_filter)
    if date_from:
        try:
            logs = logs.filter(timestamp__gte=datetime.strptime(date_from, "%Y-%m-%d"))
        except ValueError:
            pass
    if date_to:
        try:
            logs = logs.filter(timestamp__lt=datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            pass
    if search_query:
        logs = logs.filter(
            Q(description__icontains=search_query)
            | Q(object_repr__icontains=search_query)
            | Q(user__email__icontains=search_query)
        )

    return logs


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(superadmin_required)
def export_logs_excel(request):
    """Export filtered activity logs as an Excel (.xlsx) file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    logs = _filtered_logs(request)
    now = timezone.localtime(timezone.now())

    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"

    # ── Styles ────────────────────────────────────────────────────────────
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="004D1A", end_color="004D1A", fill_type="solid")
    title_font = Font(name="Calibri", bold=True, size=14, color="004D1A")
    subtitle_font = Font(name="Calibri", size=10, color="666666")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # ── Title row ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:G1")
    ws["A1"] = "KYISA Competition Management System — Activity Logs"
    ws["A1"].font = title_font

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {now.strftime('%B %d, %Y at %I:%M %p')}  |  Records: {logs.count()}"
    ws["A2"].font = subtitle_font

    # ── Column headers ────────────────────────────────────────────────────
    headers = ["#", "Date & Time", "User", "Role", "Action", "Description", "IP Address"]
    col_widths = [6, 22, 30, 20, 22, 55, 16]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # ── Data rows ─────────────────────────────────────────────────────────
    for row_idx, log in enumerate(logs[:5000], start=5):  # cap at 5 000 rows
        local_ts = timezone.localtime(log.timestamp)
        user_email = log.user.email if log.user else "System"
        user_role = log.user.get_role_display() if log.user else "-"

        values = [
            row_idx - 4,
            local_ts.strftime("%Y-%m-%d %H:%M:%S"),
            user_email,
            user_role,
            log.get_action_display(),
            log.description,
            log.ip_address or "-",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 6))

    # ── Auto-filter ───────────────────────────────────────────────────────
    last_row = 4 + min(logs.count(), 5000)
    ws.auto_filter.ref = f"A4:G{last_row}"

    # ── Write response ────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"KYISA_Activity_Logs_{now.strftime('%Y%m%d_%H%M')}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


# ══════════════════════════════════════════════════════════════════════════════
#  PDF EXPORT
# ══════════════════════════════════════════════════════════════════════════════

@login_required
@user_passes_test(superadmin_required)
def export_logs_pdf(request):
    """Export filtered activity logs as a PDF report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
    )

    logs = _filtered_logs(request)
    now = timezone.localtime(timezone.now())

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "LogTitle", parent=styles["Heading1"],
        fontSize=16, textColor=colors.HexColor("#004D1A"), spaceAfter=4,
    )
    subtitle_style = ParagraphStyle(
        "LogSubtitle", parent=styles["Normal"],
        fontSize=9, textColor=colors.gray, spaceAfter=12,
    )
    cell_style = ParagraphStyle(
        "Cell", parent=styles["Normal"], fontSize=7.5, leading=10,
    )

    elements = []

    # ── Title ─────────────────────────────────────────────────────────────
    elements.append(Paragraph("KYISA Competition Management System", title_style))
    elements.append(Paragraph("System Activity Logs Report", ParagraphStyle(
        "Sub2", parent=styles["Heading2"], fontSize=12,
        textColor=colors.HexColor("#333333"), spaceAfter=4,
    )))
    elements.append(Paragraph(
        f"Generated: {now.strftime('%B %d, %Y at %I:%M %p')}  &bull;  "
        f"Total records: {logs.count()}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 6 * mm))

    # ── Table ─────────────────────────────────────────────────────────────
    header = ["#", "Date & Time", "User", "Role", "Action", "Description", "IP"]
    data = [header]

    for idx, log in enumerate(logs[:2000], start=1):  # cap for PDF
        local_ts = timezone.localtime(log.timestamp)
        user_email = log.user.email if log.user else "System"
        user_role = log.user.get_role_display() if log.user else "-"
        desc = log.description[:120] + ("…" if len(log.description) > 120 else "")

        data.append([
            str(idx),
            local_ts.strftime("%Y-%m-%d %H:%M"),
            Paragraph(user_email, cell_style),
            user_role,
            log.get_action_display(),
            Paragraph(desc, cell_style),
            log.ip_address or "-",
        ])

    col_widths = [22, 65, 95, 65, 75, 200, 55]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    kyisa_green = colors.HexColor("#004D1A")
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), kyisa_green),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ("TOPPADDING", (0, 0), (-1, 0), 6),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        # Body
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("TOPPADDING", (0, 1), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 1), (-1, -1), 3),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        # Alternating row colours
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F7F0")]),
    ]))

    elements.append(table)

    # ── Footer note ───────────────────────────────────────────────────────
    elements.append(Spacer(1, 8 * mm))
    elements.append(Paragraph(
        f"<i>Report generated by KYISA CMS on {now.strftime('%d/%m/%Y %H:%M')}. "
        f"Confidential — for authorised personnel only.</i>",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.gray),
    ))

    doc.build(elements)
    buf.seek(0)

    filename = f"KYISA_Activity_Logs_{now.strftime('%Y%m%d_%H%M')}.pdf"
    response = HttpResponse(buf.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
